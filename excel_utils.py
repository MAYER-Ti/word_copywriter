import os
from io import BytesIO
from openpyxl import load_workbook, Workbook
import xlrd
import xlwt
from xlutils.copy import copy as xl_copy


def replace_placeholders(wb, data):
    if isinstance(wb, Workbook):
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        for key, value in data.items():
                            placeholder = f"{{{{{key}}}}}"
                            if placeholder in cell.value:
                                cell.value = cell.value.replace(placeholder, value)
    elif isinstance(wb, xlwt.Workbook):
        book = getattr(wb, "_xlrd_book")
        for idx, sheet in enumerate(book.sheets()):
            ws = wb.get_sheet(idx)
            merged = list(sheet.merged_cells)
            handled = set()
            for r in range(sheet.nrows):
                for c in range(sheet.ncols):
                    cell_value = sheet.cell_value(r, c)
                    if isinstance(cell_value, str):
                        new_value = cell_value
                        for key, value in data.items():
                            placeholder = f"{{{{{key}}}}}"
                            if placeholder in new_value:
                                new_value = new_value.replace(placeholder, value)
                        if new_value != cell_value:
                            style = xlwt.XFStyle()
                            style.xf_idx = sheet.cell_xf_index(r, c)
                            merged_range = None
                            for rng in merged:
                                rlo, rhi, clo, chi = rng
                                if rlo <= r < rhi and clo <= c < chi:
                                    merged_range = rng
                                    break
                            if merged_range and merged_range not in handled:
                                rlo, rhi, clo, chi = merged_range
                                ws.write_merge(rlo, rhi - 1, clo, chi - 1, new_value, style)
                                handled.add(merged_range)
                            else:
                                ws.write(r, c, new_value, style)
    else:
        raise TypeError("Unsupported workbook type")


def xlrd_to_openpyxl(book):
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in book.sheets():
        ws = wb.create_sheet(sheet.name)
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                value = sheet.cell_value(r, c)
                if value != "":
                    ws.cell(row=r + 1, column=c + 1, value=value)
        for rlo, rhi, clo, chi in sheet.merged_cells:
            ws.merge_cells(start_row=rlo + 1, end_row=rhi, start_column=clo + 1, end_column=chi)
    return wb


def openpyxl_to_xlwt(wb):
    book = xlwt.Workbook()
    for ws in wb.worksheets:
        sheet = book.add_sheet(ws.title)
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            for c, value in enumerate(row):
                if value is not None:
                    sheet.write(r, c, value)
        for rng in ws.merged_cells.ranges:
            rlo = rng.min_row - 1
            rhi = rng.max_row - 1
            clo = rng.min_col - 1
            chi = rng.max_col - 1
            top_left = ws.cell(rng.min_row, rng.min_col).value
            sheet.write_merge(rlo, rhi, clo, chi, top_left)
    return book


def create_document(template_bytes, ext, data, output_path):
    out_ext = os.path.splitext(output_path)[1].lower()
    if ext == ".xls":
        book = xlrd.open_workbook(file_contents=template_bytes, formatting_info=True)
        if out_ext == ".xls":
            wb = xl_copy(book)
            wb._xlrd_book = book
            replace_placeholders(wb, data)
            wb.save(output_path)
        else:
            wb = xlrd_to_openpyxl(book)
            replace_placeholders(wb, data)
            for ws in wb.worksheets:
                ws.sheet_state = "visible"
            wb.save(output_path)
    else:
        wb = load_workbook(BytesIO(template_bytes))
        replace_placeholders(wb, data)
        for ws in wb.worksheets:
            ws.sheet_state = "visible"
        if out_ext == ".xls":
            book = openpyxl_to_xlwt(wb)
            book.save(output_path)
        else:
            wb.save(output_path)
